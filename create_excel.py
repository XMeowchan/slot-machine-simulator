import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 现有契约（绿色背景 - 已完成）
existing_contracts = [
    (1, "启动资金", "StartingFund", "根据当前阶段，获得一份神秘奖励"),
    (2, "飞升", "Ascend", "战斗开始{0}秒后，所有友军获得大幅属性增强"),
    (3, "捆绑出售", "TwoStarForSale", "商店有几率直接刷出二星棋子"),
    (4, "潘多拉棋盘", "PandoraChessboa", "将当前上阵的所有棋子变形为质量高一级的同星级随机棋子"),
    (5, "玻璃大炮", "GlassCannon", "处于最后一排的棋子最大生命值仅为70%，攻击力提升至150%"),
    (6, "性价比之选", "BlueIsGood", "所有棋子根据品质获得属性加成：低星棋子最高+{0}，高星棋子最低+{1}"),
    (7, "稳定货源", "ThirdRollBlue", "每第三次刷新，商店中刷新全部为紫色品质的棋子"),
    (8, "带队大哥", "BigBoss", "每场伤害最高的棋子，在下一回合开始时获得全属性增益"),
    (9, "最后的波纹", "LastDance", "友军单位死亡时，对附近的友军造成基于自身最大生命值50%的治疗"),
    (10, "守财奴", "SaveMoney", "每持有{0}金币可增加{1}%的胜利赏金，最多增加{2}%"),
    (11, "不要小看我们的羁绊", "RaceMaster", "所有需要{0}个及以上羁绊类型触发的效果，羁绊类型需求量-{1}"),
    (12, "i人", "IHuman", "商店中有更高概率刷到自己已有的棋子"),
    (13, "e人", "EHuman", "商店中有更高概率刷到自己没有的棋子"),
    (14, "回炉重造", "ItemMaster", "连续{0}回合战斗都没有被装备的道具会进化成一个品质+{1}的道具。相同道具同时存在时，进化进度加快（最高品质：5）"),
    (15, "大复制术", "CopyItem", "获得道具时，额外获得一个相同的道具"),
    (16, "MVP!", "MVP", "战斗结束后，有概率再获得一个本场战斗伤害最高的棋子。获得棋子的星级为1。"),
    (17, "刷刷刷刷刷", "RefreshShop", "刷新商店的价格降低至{0}金币"),
    (18, "老朋友", "SameChess", "已拥有的棋子在商店中的价格降低{0}%"),
    (19, "草莽英雄", "LowChessPower", "战斗中，若友方场上只剩下1品质棋子，其伤害和生命值总量增加{0}%"),
    (20, "孤勇者", "SoloChess", "战斗中，若友方场上只剩下1个棋子，其伤害和生命值总量增加{0}%"),
    (21, "原来你也......", "SameRace", "棋子的随机羁绊会是你上一次获得的随机羁绊。若无获得过随机羁绊或无法获得该羁绊，则随机抽选。"),
]

# 新增契约
new_contracts = [
    (22, "你快乐吗？", "HappyMaker", "快乐使者棋子攻击敌人时有概率将其变为友方棋子", "娱乐向控制流"),
    (23, "VIP门票", "VIPPass", "每回合基于存款获得免费刷新次数，每拥有{0}枚金币，获得{1}次免费刷新次数", "经济运营流"),
    (24, "神器锻造炉", "ArtifactForge", "{0}个战斗回合后，获得一件神器（狙击镜、冻结之锤、不死藤甲、狂战士之斧、真·防身泪滴）", "延迟奖励流"),
    (25, "变形重组器", "Recombinator", "失去当前的所有棋子，根据棋子的总品质和星级获得一个高阶棋子", "高风险高收益"),
    (26, "啦啦队", "CheerSquad", "你在场的每个白色和绿色品质棋子都会为你的紫色品质及以上的棋子提供伤害和攻速增幅", "低星辅助高星"),
    (27, "训练假人桩", "TrainingDummy", "获得一个训练假人桩，它无法攻击，但会提供{0}种随机羁绊；假人桩死亡{1}次后，随机获得{2}个珍贵道具", "功能性羁绊工具"),
    (28, "风水轮流转", "RotatingFortune", "你伤害最低的棋子，在下一回合开始时将获得全属性增益", "弱者保护机制"),
    (29, "忍耐是一种美德", "Patience", "接下来三个回合无法购买任何棋子，三个回合后，获得一个高品质棋子", "延迟满足"),
    (30, "亡羊补牢", "BetterLate", "战斗失败时，下回合商店首次刷新必定出现至少一个紫色品质棋子", "连败保护机制"),
    (31, "趁热打铁", "StrikeWhileHot", "连胜时，每多一场连胜，所有棋子攻击力额外增加{0}%（最多{1}%）", "连胜奖励放大"),
    (32, "背水一战", "LastStand", "当场上只剩1个棋子时，该棋子获得{0}%生命偷取和{1}%攻击速度", "与孤勇者形成combo"),
    (33, "以战养战", "WarEconomy", "每击杀一个敌方单位，获得{0}金币", "击杀奖励经济流"),
    (34, "装备大师", "GearMaster", "棋子装备道具后，额外获得该道具{0}%的属性加成", "装备强化流"),
    (35, "人海战术", "ZergRush", "人口上限+{0}，但所有棋子最大生命值降低{1}%", "数量vs质量的取舍"),
    (36, "精英小队", "EliteSquad", "人口上限-{0}，但所有棋子攻击力和生命值增加{1}%", "与人海战术对立"),
    (37, "幸运抽奖", "LuckyDraw", "每回合开始时，有{0}%概率获得一个随机道具", "随机奖励流"),
    (38, "血祭", "BloodSacrifice", "每场战斗开始时，损失{0}%当前生命值，所有棋子伤害增加{1}%", "高风险高回报"),
    (39, "时光倒流", "TimeRewind", "每{0}回合，可以保留一次商店刷新结果到下一回合", "商店运营策略"),
    (40, "羁绊共鸣", "BondResonance", "激活的羁绊数量每增加1个，所有棋子获得{0}%全属性加成", "多羁绊流核心"),
    (41, "星之力", "StarPower", "三星棋子额外获得{0}%暴击率和{1}%暴击伤害", "追三星流"),
    (42, "前期压制", "EarlyPressure", "前{0}回合，所有友军攻击速度增加{1}%", "前期强势流"),
]

# 创建工作簿
wb = Workbook()

# 删除默认工作表
wb.remove(wb.active)

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
existing_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
new_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 黄色
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== Sheet 1: 所有契约总览 =====
ws1 = wb.create_sheet("所有契约总览")

# 写入标题
headers = ["序号", "契约名称", "英文代码", "效果描述", "状态"]
ws1.append(headers)

# 设置表头样式
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# 写入现有契约
for row in existing_contracts:
    ws1.append([row[0], row[1], row[2], row[3], "已完成"])
    row_num = ws1.max_row
    for col in range(1, 6):
        cell = ws1.cell(row=row_num, column=col)
        cell.fill = existing_fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# 写入新契约
for row in new_contracts:
    ws1.append([row[0], row[1], row[2], row[3], "新设计"])
    row_num = ws1.max_row
    for col in range(1, 6):
        cell = ws1.cell(row=row_num, column=col)
        cell.fill = new_fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# 调整列宽
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 70
ws1.column_dimensions['E'].width = 12

# ===== Sheet 2: 现有契约（已完成） =====
ws2 = wb.create_sheet("现有契约-已完成")
ws2.append(["序号", "契约名称", "英文代码", "效果描述"])

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for row in existing_contracts:
    ws2.append([row[0], row[1], row[2], row[3]])
    row_num = ws2.max_row
    for col in range(1, 5):
        cell = ws2.cell(row=row_num, column=col)
        cell.fill = existing_fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 80

# ===== Sheet 3: 新契约设计 =====
ws3 = wb.create_sheet("新契约设计")
ws3.append(["序号", "契约名称", "英文代码", "效果描述", "设计思路"])

for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for row in new_contracts:
    ws3.append([row[0], row[1], row[2], row[3], row[4]])
    row_num = ws3.max_row
    for col in range(1, 6):
        cell = ws3.cell(row=row_num, column=col)
        cell.fill = new_fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 65
ws3.column_dimensions['E'].width = 20

# ===== Sheet 4: 新增契约分类 =====
ws4 = wb.create_sheet("新增契约分类")
ws4.append(["类型", "契约名称"])

for cell in ws4[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

categories = [
    ("经济运营", ["VIP门票", "以战养战", "趁热打铁", "时光倒流"]),
    ("战斗强化", ["趁热打铁", "亡羊补牢", "背水一战", "血祭", "前期压制"]),
    ("装备道具", ["神器锻造炉", "装备大师", "大复制术", "回炉重造", "幸运抽奖"]),
    ("阵容策略", ["变形重组器", "啦啦队", "人海战术", "精英小队", "羁绊共鸣"]),
    ("风险控制", ["你快乐吗？", "变形重组器", "忍耐是一种美德", "亡羊补牢", "血祭"]),
    ("特色机制", ["你快乐吗？", "训练假人桩", "风水轮流转", "星之力", "原来你也......"]),
]

for cat_name, items in categories:
    for item in items:
        ws4.append([cat_name, item])
        row_num = ws4.max_row
        for col in range(1, 3):
            cell = ws4.cell(row=row_num, column=col)
            cell.fill = new_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 25

# 保存文件
output_path = r"C:\Users\Administrator\Desktop\战魂铭人自走棋-契约设计.xlsx"
wb.save(output_path)
print(f"Excel文件已保存至: {output_path}")
